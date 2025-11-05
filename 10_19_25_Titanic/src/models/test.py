import xlwings as xw
import pandas as pd
from sklearn.metrics import ConfusionMatrixDisplay, classification_report, RocCurveDisplay
from src.data.load_data import load_dataset
from src.utils.helpers import plot_confusion_matrix, plot_roc, classification_report_df
from src.visualization.eda_plot import evaluate_plots, test_performance, make_plots
from pathlib import Path
from joblib import dump
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from io import BytesIO
import json


def test_all_models(models_dict,X_test) -> pd.DataFrame: 
    predictions = dict()
    for key in models_dict.keys():
        preds = models_dict[key].predict(X_test)
        predictions[key] = preds
        dump(models_dict[key],f'reports/model_outputs/dumped_models/{key}_trained.pkl')
    return pd.DataFrame(predictions)

def create_report(predictions,X_train, y_train, y_test,trained_models, path=None)->None:
    if path is None:
        path = 'reports/model_outputs/results/final_report.xlsx'
    eda_plots = make_plots(X_train, y_train, ['Embarked', 'Sex', 'Pclass'],
          ['Family_size', 'Fare', 'Parch', 'SibSp']);
    cm,roc = evaluate_plots(predictions,y_test)
    metrics_plot, reports_df = test_performance(predictions, y_test)
    config = {}
    for key, model in trained_models.items():
        config[key]= model.named_steps['model'].get_params()
    with open('reports/model_outputs/results/model_configs.json', 'w') as f:
        json.dump(config, f, indent=4)
    with pd.ExcelWriter(path) as writer:
        for i, df in reports_df.items():
            sheet_name=f'metrics_{i}'
            df.to_excel(writer, sheet_name=sheet_name, index=True, float_format="%.3f")

    wb = load_workbook(path)
        
    ws_eda = wb.create_sheet("EDA")
    row = 1
    for fig in eda_plots:
        buf=BytesIO()
        fig.savefig(buf, format='jpg', bbox_inches='tight')
        buf.seek(0)
        img = Image(buf)
        img.width, img.height = 800,400
        ws_eda.add_image(img, f'A{row}')
        row += 25
        plt.close(fig)
    ws_metrics = wb.create_sheet("Metrics Plots")
    row = 1
    for i, fig in enumerate([cm, roc, metrics_plot], start=1):
        buf=BytesIO()
        fig.savefig(buf, format='jpg', bbox_inches='tight')
        buf.seek(0)
        img = Image(buf)
        img.width, img.height = 800,400
        ws_metrics.add_image(img, f'A{row}')
        row += 25
        plt.close(fig)

    wb.save(path)
    print(f'Report was saved to: {path}')
